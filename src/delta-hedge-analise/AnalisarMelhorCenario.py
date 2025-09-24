import sys
import os
import pandas as pd
from datetime import datetime
import sqlite3
from openpyxl.utils import get_column_letter

def carregar_dados_excel(arquivo_excel):
    """
    Carrega os dados de um arquivo Excel de simulação.
    """
    try:
        df = pd.read_excel(arquivo_excel)
        print(f"Carregados {len(df)} registros de {arquivo_excel}")
        return df
    except Exception as e:
        print(f"Erro ao carregar {arquivo_excel}: {str(e)}")
        return None

def limpar_valores_monetarios(valor):
    """
    Remove formatação monetária e converte para float.
    """
    if pd.isna(valor) or valor == 'N/A':
        return None
    
    if isinstance(valor, str):
        # Remove "R$ " e converte vírgula para ponto
        valor_limpo = valor.replace('R$ ', '').replace(',', '.')
        try:
            return float(valor_limpo)
        except ValueError:
            return None
    
    return float(valor)

def limpar_valores_delta(valor):
    """
    Remove formatação de delta e converte para float.
    """
    if pd.isna(valor) or valor == 'N/A':
        return None
    
    if isinstance(valor, str):
        try:
            return float(valor)
        except ValueError:
            return None
    
    return float(valor)

def obter_precos_petrobras(conn, data_inicio, data_termino):
    """
    Obtém os preços da Petrobras no início e fim da simulação.
    """
    try:
        cursor = conn.cursor()
        
        # Busca preço no início
        cursor.execute("""
            SELECT abertura FROM HIST_ATIVO 
            WHERE id_ativo = 1 AND data = ?
        """, (data_inicio,))
        preco_inicio = cursor.fetchone()
        
        # Busca preço no fim
        cursor.execute("""
            SELECT abertura FROM HIST_ATIVO 
            WHERE id_ativo = 1 AND data = ?
        """, (data_termino,))
        preco_fim = cursor.fetchone()
        
        return (preco_inicio[0] if preco_inicio else None, 
                preco_fim[0] if preco_fim else None)
                
    except Exception as e:
        print(f"Erro ao buscar preços: {str(e)}")
        return (None, None)

def main():
    # Caminhos dos arquivos
    arquivo_delta = 'dados/SimulacaoPeloDelta.xlsx'
    arquivo_dia = 'dados/SimulacaoPeloDia.xlsx'
    arquivo_lote = 'dados/SimulacaoPeloLote.xlsx'
    arquivo_saida = 'dados/MelhorCenario.xlsx'
    
    # Conecta ao banco de dados
    conn = sqlite3.connect('banco/mercado_opcoes.db')
    
    try:
        print("Carregando dados dos arquivos Excel...")
        
        # Carrega os três arquivos
        df_delta = carregar_dados_excel(arquivo_delta)
        df_dia = carregar_dados_excel(arquivo_dia)
        df_lote = carregar_dados_excel(arquivo_lote)
        
        if df_delta is None or df_dia is None or df_lote is None:
            print("Erro ao carregar um ou mais arquivos Excel.")
            return
        
        # Adiciona coluna de tipo de estratégia
        df_delta['Ajuste'] = 'Delta'
        df_dia['Ajuste'] = 'Dias'
        df_lote['Ajuste'] = 'Lote'
        
        # Adiciona coluna com o valor do parâmetro de ajuste
        df_delta['Valor'] = df_delta['Ajuste.Delta']
        df_dia['Valor'] = df_dia['Freq.Ajuste']
        df_lote['Valor'] = df_lote['Limite Lote']
        
        # Remove colunas específicas de cada estratégia
        df_delta = df_delta.drop(columns=['Ajuste.Delta'])
        df_dia = df_dia.drop(columns=['Freq.Ajuste'])
        df_lote = df_lote.drop(columns=['Limite Lote'])
        
        # Combina os três DataFrames
        df_combinado = pd.concat([df_delta, df_dia, df_lote], ignore_index=True)
        print(f"Total de registros combinados: {len(df_combinado)}")
        
        # Limpa valores monetários para comparação
        df_combinado['Saldo Final Limpo'] = df_combinado['Saldo Final'].apply(limpar_valores_monetarios)
        
        # Agrupa por simulação e encontra o melhor cenário
        dados_melhor_cenario = []
        
        # Agrupa por opção, vencimento e período para identificar simulações únicas
        for (opcao, vencimento, inicio, termino), grupo in df_combinado.groupby(['Opção', 'Vencimento', 'Início', 'Término']):
            print(f"\nAnalisando simulação: {opcao} - {vencimento} ({inicio} a {termino})")
            
            # Encontra o melhor cenário (maior saldo final)
            melhor_idx = grupo['Saldo Final Limpo'].idxmax()
            melhor_cenario = grupo.loc[melhor_idx]
            
            print(f"  Melhor estratégia: {melhor_cenario['Ajuste']} (Valor: {melhor_cenario['Valor']})")
            print(f"  Saldo final: {melhor_cenario['Saldo Final']}")
            
            # Obtém preços da Petrobras
            preco_inicio, preco_fim = obter_precos_petrobras(conn, inicio, termino)
            
            dados_melhor_cenario.append({
                'Opção': melhor_cenario['Opção'],
                'Vencimento': melhor_cenario['Vencimento'],
                'Strike': melhor_cenario['Strike'],
                'Início': melhor_cenario['Início'],
                'Término': melhor_cenario['Término'],
                'PETR-Início': preco_inicio,
                'PETR-Término': preco_fim,
                'Δ Inicio': melhor_cenario['Δ Inicio'],
                'Δ Fim': melhor_cenario['Δ Fim'],
                'Simulação': melhor_cenario['Simulação'],
                'Ajuste': melhor_cenario['Ajuste'],
                'Valor': melhor_cenario['Valor'],
                '# Pregões Vol.': melhor_cenario['# Pregões Vol.'],
                '# Ajustes': melhor_cenario['# Ajustes'],
                'Saldo Final': melhor_cenario['Saldo Final'],
                'Melhor Saldo': melhor_cenario['Melhor Saldo'],
                'Data Melhor Saldo': melhor_cenario['Data Melhor Saldo']
            })
        
        # Cria DataFrame final
        df_final = pd.DataFrame(dados_melhor_cenario)
        
        # Formatação das colunas
        for col in ['Strike', 'PETR-Início', 'PETR-Término', 'Melhor Saldo', 'Saldo Final']:
            df_final[col] = df_final[col].apply(lambda x: 
                f'R$ {float(x):.2f}' if pd.notnull(x) and x != 'N/A' and not isinstance(x, str) else 
                x if isinstance(x, str) and x.startswith('R$') else 
                f'R$ {x:.2f}' if pd.notnull(x) and x != 'N/A' else 'N/A'
            )
        for col in ['Δ Inicio', 'Δ Fim']:
            df_final[col] = df_final[col].apply(lambda x: 
                f'{float(x):.4f}' if pd.notnull(x) and x != 'N/A' and not isinstance(x, str) else 
                x if isinstance(x, str) and not x.startswith('R$') else 
                f'{x:.4f}' if pd.notnull(x) and x != 'N/A' else 'N/A'
            )
        
        # Salva no Excel
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name='MelhorCenario', index=False)
            worksheet = writer.sheets['MelhorCenario']
            
            # Ajusta a largura das colunas
            for col_idx, col_name in enumerate(df_final.columns, 1):
                try:
                    max_len = df_final[col_name].astype(str).map(len).max()
                    header_len = len(col_name)
                    column_width = max(max_len, header_len) + 2
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width
                except (ValueError, TypeError):
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 15

        print(f"\nArquivo consolidado salvo em: {arquivo_saida}")
        print(f"Total de simulações processadas: {len(dados_melhor_cenario)}")
        
        # Resumo por tipo de ajuste
        print("\nResumo por tipo de ajuste:")
        print("=" * 50)
        resumo_tipo = df_final['Ajuste'].value_counts()
        for tipo, count in resumo_tipo.items():
            print(f"{tipo}: {count} simulações")
        
        # Resumo dos valores de ajuste
        print("\nResumo dos valores de ajuste:")
        print("=" * 50)
        for tipo in ['Delta', 'Dias', 'Lote']:
            valores = df_final[df_final['Ajuste'] == tipo]['Valor'].value_counts()
            print(f"\n{tipo}:")
            for valor, count in valores.items():
                print(f"  {valor}: {count} simulações")
        
    except Exception as e:
        print(f"Erro durante a análise: {str(e)}")
        import traceback
        traceback.print_exc()
    
    finally:
        conn.close()

if __name__ == "__main__":
    main() 