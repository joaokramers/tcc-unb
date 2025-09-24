import sys
import os
import re
import pandas as pd
from datetime import datetime
import sqlite3
from openpyxl.utils import get_column_letter

def extrair_dados_simulacao(arquivo_txt):
    """
    Extrai os dados das simulações do arquivo de texto.
    Retorna: (simulacoes_melhor_cenario, todos_cenarios)
    """
    simulacoes = []
    todos_cenarios = []
    
    with open(arquivo_txt, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    
    print(f"Tamanho do arquivo: {len(conteudo)} caracteres")
    
    # Divide o conteúdo em seções por simulação
    secoes = conteudo.split("EXECUTANDO CENÁRIOS PARA SIMULAÇÃO ID")
    print(f"Número de seções encontradas: {len(secoes)}")
    
    for i, secao in enumerate(secoes[1:], 1):  # Pula a primeira seção (cabeçalho)
        try:
            print(f"\nProcessando seção {i}...")
            
            # Extrai ID da simulação
            id_match = re.search(r'(\d+)', secao)
            if not id_match:
                print("ID da simulação não encontrado")
                continue
            id_simulacao = int(id_match.group(1))
            print(f"ID da simulação: {id_simulacao}")
            
            # Extrai dados da opção
            ticker_match = re.search(r'Ticker: (\w+)', secao)
            strike_match = re.search(r'Strike: R\$ ([\d.]+)', secao)
            vencimento_match = re.search(r'Vencimento: (\d{4}-\d{2}-\d{2})', secao)
            
            if not all([ticker_match, strike_match, vencimento_match]):
                print("Dados da opção incompletos")
                continue
                
            ticker = ticker_match.group(1)
            strike = float(strike_match.group(1))
            vencimento = vencimento_match.group(1)
            print(f"Opção: {ticker}, Strike: {strike}, Vencimento: {vencimento}")
            
            # Extrai período da simulação
            periodo_match = re.search(r'Período: (\d{4}-\d{2}-\d{2}) até (\d{4}-\d{2}-\d{2})', secao)
            if not periodo_match:
                print("Período da simulação não encontrado")
                continue
                
            data_inicio = periodo_match.group(1)
            data_termino = periodo_match.group(2)
            print(f"Período: {data_inicio} até {data_termino}")
            
            # Extrai dados dos cenários
            cenarios = []
            cenarios_secoes = secao.split("CENÁRIO:")
            print(f"Número de cenários encontrados: {len(cenarios_secoes) - 1}")
            
            for j, cenario_sec in enumerate(cenarios_secoes[1:], 1):  # Pula a primeira seção
                print(f"  Processando cenário {j}...")
                
                # Extrai parâmetros do cenário (adaptado para limite de lote)
                limite_match = re.search(r'Limite Lote = (\d+)', cenario_sec)
                volatilidade_match = re.search(r'Pregões Volatilidade = (\d+)', cenario_sec)
                
                if not all([limite_match, volatilidade_match]):
                    print(f"    Parâmetros do cenário {j} não encontrados")
                    continue
                    
                limite_lote = int(limite_match.group(1))
                pregoes_volatilidade = int(volatilidade_match.group(1))
                print(f"    Parâmetros: Limite Lote={limite_lote}, Vol={pregoes_volatilidade}")
                
                # Extrai número de ajustes
                # Novo método: conta quantas vezes aparece 'True' na coluna Ajuste da tabela
                # A coluna Ajuste aparece antes do 'Saldo Real' e depois do 'Saldo Acumulado'
                # Exemplo de linha:
                # 2025-04-02 R$ 36.18 R$ 7.40 0.9986 ... R$ -28730.54    R$ -28730.54    True     R$ 0.00
                ajustes_true = re.findall(r'\s(True)\s+R\$\s*-?[\d.,]+', cenario_sec)
                num_ajustes = len(ajustes_true)
                print(f"    Número de ajustes (contando True na tabela): {num_ajustes}")
                
                # Extrai deltas inicial e final
                # O padrão \b(\d\.\d{4})\b busca por um número com 1 dígito, ponto, e 4 dígitos (formato do delta)
                delta_matches = re.findall(r'\b(\d\.\d{4})\b', cenario_sec)
                delta_inicial = float(delta_matches[0]) if len(delta_matches) >= 1 else None
                delta_final = float(delta_matches[-1]) if len(delta_matches) >= 1 else None
                
                # Extrai saldo final - tenta múltiplas abordagens
                saldo_final = 0.0
                
                # Abordagem 1: Último valor de Saldo Real na tabela
                saldo_real_matches = re.findall(r'Saldo Real\s+R\$ ([\d.-]+)', cenario_sec)
                if saldo_real_matches:
                    saldo_final = float(saldo_real_matches[-1])
                    print(f"    Saldo final (último da tabela): R$ {saldo_final:.2f}")
                else:
                    # Abordagem 2: Busca por "Saldo Real" seguido de valor
                    saldo_match = re.search(r'Saldo Real\s+R\$ ([\d.-]+)', cenario_sec)
                    if saldo_match:
                        saldo_final = float(saldo_match.group(1))
                        print(f"    Saldo final (primeiro encontrado): R$ {saldo_final:.2f}")
                    else:
                        # Abordagem 3: Busca por valores monetários no final
                        valores_monetarios = re.findall(r'R\$ ([\d.-]+)', cenario_sec)
                        if valores_monetarios:
                            saldo_final = float(valores_monetarios[-1])
                            print(f"    Saldo final (último valor monetário): R$ {saldo_final:.2f}")
                        else:
                            print(f"    Saldo final não encontrado")
                
                # Extrai melhor saldo e data do melhor saldo
                melhor_saldo = -float('inf')
                data_melhor_saldo = None
                rows_matches = re.findall(r'^(\d{4}-\d{2}-\d{2}).*?(?:True|False)\s+R\$\s*(-?[\d.]+)\s*$', cenario_sec, re.MULTILINE)

                if rows_matches:
                    data_saldo_pairs = [(match[0], float(match[1])) for match in rows_matches]
                    if data_saldo_pairs:
                        data_melhor_saldo, melhor_saldo = max(data_saldo_pairs, key=lambda item: item[1])
                
                cenario_data = {
                    'limite_lote': limite_lote,
                    'pregoes_volatilidade': pregoes_volatilidade,
                    'num_ajustes': num_ajustes,
                    'saldo_final': saldo_final,
                    'delta_inicial': delta_inicial,
                    'delta_final': delta_final,
                    'melhor_saldo': melhor_saldo if melhor_saldo != -float('inf') else None,
                    'data_melhor_saldo': data_melhor_saldo
                }
                cenarios.append(cenario_data)
                
                # Adiciona todos os cenários para a aba "Todos"
                todos_cenarios.append({
                    'id_simulacao': id_simulacao,
                    'ticker': ticker,
                    'strike': strike,
                    'vencimento': vencimento,
                    'data_inicio': data_inicio,
                    'data_termino': data_termino,
                    **cenario_data
                })
            
            # Pega o melhor cenário (maior saldo final) para a aba principal
            if cenarios:
                melhor_cenario = max(cenarios, key=lambda x: x['saldo_final'])
                print(f"Melhor cenário: Limite Lote={melhor_cenario['limite_lote']}, Vol={melhor_cenario['pregoes_volatilidade']}, Saldo=R$ {melhor_cenario['saldo_final']:.2f}")
                
                simulacoes.append({
                    'id_simulacao': id_simulacao,
                    'ticker': ticker,
                    'strike': strike,
                    'vencimento': vencimento,
                    'data_inicio': data_inicio,
                    'data_termino': data_termino,
                    'limite_lote': melhor_cenario['limite_lote'],
                    'pregoes_volatilidade': melhor_cenario['pregoes_volatilidade'],
                    'num_ajustes': melhor_cenario['num_ajustes'],
                    'saldo_final': melhor_cenario['saldo_final'],
                    'delta_inicial': melhor_cenario['delta_inicial'],
                    'delta_final': melhor_cenario['delta_final'],
                    'melhor_saldo': melhor_cenario['melhor_saldo'],
                    'data_melhor_saldo': melhor_cenario['data_melhor_saldo']
                })
            else:
                print("Nenhum cenário válido encontrado")
                
        except Exception as e:
            print(f"Erro ao processar simulação {i}: {str(e)}")
            continue
    
    return simulacoes, todos_cenarios

def obter_precos_petrobras(conn, data_inicio, data_termino):
    """
    Obtém os preços da Petrobras no início e fim da simulação.
    """
    try:
        cursor = conn.cursor()
        
        # Busca preço no início
        cursor.execute("""
            SELECT abertura FROM HIST_ATIVO 
            WHERE id_ativo = 1 AND data = ?
        """, (data_inicio,))
        preco_inicio = cursor.fetchone()
        
        # Busca preço no fim (fechamento do último dia)
        cursor.execute("""
            SELECT fechamento FROM HIST_ATIVO 
            WHERE id_ativo = 1 AND data = ?
        """, (data_termino,))
        preco_fim = cursor.fetchone()
        
        return (preco_inicio[0] if preco_inicio else None, 
                preco_fim[0] if preco_fim else None)
                
    except Exception as e:
        print(f"Erro ao buscar preços: {str(e)}")
        return (None, None)

def determinar_situacao_delta(delta_inicial, delta_final):
    """
    Determina a situação da opção baseada nos valores de Delta.
    
    Classificação:
    - 0.00 a 0.30: Opção fora do dinheiro (OTM)
    - 0.31 a 0.70: Opção no dinheiro (ATM)
    - 0.71 a 1.00: Opção dentro do dinheiro (ITM)
    """
    def classificar_delta(delta):
        if delta is None:
            return "N/A"
        elif delta <= 0.30:
            return "OTM"
        elif delta <= 0.70:
            return "ATM"
        else:
            return "ITM"
    
    situacao_inicio = classificar_delta(delta_inicial)
    situacao_fim = classificar_delta(delta_final)
    
    return f"{situacao_inicio} → {situacao_fim}"

def main():
    # Caminhos dos arquivos
    arquivo_txt = 'dados/SimulacaoPeloLote.txt'
    arquivo_excel = 'dados/SimulacaoPeloLote.xlsx'
    
    # Conecta ao banco de dados
    conn = sqlite3.connect('banco/mercado_opcoes.db')
    
    try:
        print("Analisando arquivo de simulações...")
        simulacoes, todos_cenarios = extrair_dados_simulacao(arquivo_txt)
        
        if not simulacoes:
            print("Nenhuma simulação encontrada no arquivo.")
            return
        
        print(f"Encontradas {len(simulacoes)} simulações para sumarizar.")
        print(f"Total de cenários (todas as combinações): {len(todos_cenarios)}")
        
        # Cria DataFrame com os dados
        dados = []
        
        for sim in simulacoes:
            # Obtém preços da Petrobras
            preco_inicio, preco_fim = obter_precos_petrobras(
                conn, sim['data_inicio'], sim['data_termino']
            )
            
            # Obtém valor da opção no primeiro dia
            cursor = conn.cursor()
            cursor.execute("""
                SELECT h.fechamento
                FROM HIST_OPCAO h
                JOIN OPCAO o ON o.id = h.id_opcao
                WHERE o.ticker = ? AND h.data = ?
                ORDER BY h.data ASC
                LIMIT 1
            """, (sim['ticker'], sim['data_inicio']))
            
            valor_opcao_result = cursor.fetchone()
            valor_opcao = valor_opcao_result[0] if valor_opcao_result else None
            
            # Determina situação da opção baseada no Delta
            situacao = determinar_situacao_delta(sim['delta_inicial'], sim['delta_final'])
            
            dados.append({
                'Opção': sim['ticker'],
                'Vencimento': sim['vencimento'],
                'Strike': sim['strike'],
                'Preço': valor_opcao,
                'Início': sim['data_inicio'],
                'Término': sim['data_termino'],
                'PETR-Início': preco_inicio,
                'PETR-Término': preco_fim,
                'Δ Inicio': sim['delta_inicial'],
                'Δ Fim': sim['delta_final'],
                'Simulação': situacao,
                'Limite Lote': sim['limite_lote'],
                '# Pregões Vol.': sim['pregoes_volatilidade'],
                '# Ajustes': sim['num_ajustes'],
                'Saldo Final': sim['saldo_final'],
                'Melhor Saldo': sim['melhor_saldo'],
                'Data Melhor Saldo': sim['data_melhor_saldo']
            })
        
        # Cria DataFrame
        df = pd.DataFrame(dados)
        
        # Formatação das colunas
        for col in ['Strike', 'Preço', 'PETR-Início', 'PETR-Término', 'Melhor Saldo', 'Saldo Final']:
            df[col] = df[col].apply(lambda x: f'R$ {x:.2f}' if pd.notnull(x) else 'N/A')
        for col in ['Δ Inicio', 'Δ Fim']:
            df[col] = df[col].apply(lambda x: f'{x:.4f}' if pd.notnull(x) else 'N/A')
        
        # Cria DataFrame para todos os cenários
        dados_todos = []
        
        for cenario in todos_cenarios:
            # Obtém preços da Petrobras
            preco_inicio, preco_fim = obter_precos_petrobras(
                conn, cenario['data_inicio'], cenario['data_termino']
            )
            
            # Obtém valor da opção no primeiro dia
            cursor = conn.cursor()
            cursor.execute("""
                SELECT h.fechamento
                FROM HIST_OPCAO h
                JOIN OPCAO o ON o.id = h.id_opcao
                WHERE o.ticker = ? AND h.data = ?
                ORDER BY h.data ASC
                LIMIT 1
            """, (cenario['ticker'], cenario['data_inicio']))
            
            valor_opcao_result = cursor.fetchone()
            valor_opcao = valor_opcao_result[0] if valor_opcao_result else None
            
            # Determina situação da opção baseada no Delta
            situacao = determinar_situacao_delta(cenario['delta_inicial'], cenario['delta_final'])
            
            dados_todos.append({
                'Opção': cenario['ticker'],
                'Vencimento': cenario['vencimento'],
                'Strike': cenario['strike'],
                'Preço': valor_opcao,
                'Início': cenario['data_inicio'],
                'Término': cenario['data_termino'],
                'PETR-Início': preco_inicio,
                'PETR-Término': preco_fim,
                'Δ Inicio': cenario['delta_inicial'],
                'Δ Fim': cenario['delta_final'],
                'Simulação': situacao,
                'Limite Lote': cenario['limite_lote'],
                '# Pregões Vol.': cenario['pregoes_volatilidade'],
                '# Ajustes': cenario['num_ajustes'],
                'Saldo Final': cenario['saldo_final'],
                'Melhor Saldo': cenario['melhor_saldo'],
                'Data Melhor Saldo': cenario['data_melhor_saldo']
            })
        
        # Cria DataFrame para todos os cenários
        df_todos = pd.DataFrame(dados_todos)
        
        # Formatação das colunas para todos os cenários
        for col in ['Strike', 'Preço', 'PETR-Início', 'PETR-Término', 'Melhor Saldo', 'Saldo Final']:
            df_todos[col] = df_todos[col].apply(lambda x: f'R$ {x:.2f}' if pd.notnull(x) else 'N/A')
        for col in ['Δ Inicio', 'Δ Fim']:
            df_todos[col] = df_todos[col].apply(lambda x: f'{x:.4f}' if pd.notnull(x) else 'N/A')
        
        # Salva no Excel
        with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
            # Aba principal (melhor cenário de cada opção)
            df.to_excel(writer, sheet_name='SimulacaoPeloLote', index=False)
            worksheet = writer.sheets['SimulacaoPeloLote']
            
            # Ajusta a largura das colunas de forma mais robusta
            for col_idx, col_name in enumerate(df.columns, 1):
                try:
                    # Calcula o tamanho máximo do conteúdo da coluna
                    max_len = df[col_name].astype(str).map(len).max()
                    # Compara com o tamanho do cabeçalho
                    header_len = len(col_name)
                    # Usa o maior valor + 2 de buffer
                    column_width = max(max_len, header_len) + 2
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width
                except (ValueError, TypeError):
                    # Se houver erro (ex: coluna vazia), usa um valor padrão
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # Aba "Todos" (todas as combinações)
            df_todos.to_excel(writer, sheet_name='Todos', index=False)
            worksheet_todos = writer.sheets['Todos']
            
            # Ajusta a largura das colunas para a aba "Todos"
            for col_idx, col_name in enumerate(df_todos.columns, 1):
                try:
                    # Calcula o tamanho máximo do conteúdo da coluna
                    max_len = df_todos[col_name].astype(str).map(len).max()
                    # Compara com o tamanho do cabeçalho
                    header_len = len(col_name)
                    # Usa o maior valor + 2 de buffer
                    column_width = max(max_len, header_len) + 2
                    worksheet_todos.column_dimensions[get_column_letter(col_idx)].width = column_width
                except (ValueError, TypeError):
                    # Se houver erro (ex: coluna vazia), usa um valor padrão
                    worksheet_todos.column_dimensions[get_column_letter(col_idx)].width = 15

        print(f"Tabela salva em: {arquivo_excel}")
        print(f"Total de simulações processadas: {len(simulacoes)}")
        print(f"Total de cenários na aba 'Todos': {len(todos_cenarios)}")
        print(f"Arquivo Excel criado com 2 abas:")
        print(f"  - 'SimulacaoPeloLote': Melhor cenário de cada opção ({len(simulacoes)} registros)")
        print(f"  - 'Todos': Todas as combinações de limite de lote e volatilidade ({len(todos_cenarios)} registros)")
        
        # Resumo
        saldos_finais = [s['saldo_final'] for s in simulacoes if s['saldo_final'] is not None]
        if saldos_finais:
            print("\nResumo das simulações:")
            print("=" * 80)
            print(f"Melhor saldo final: R$ {max(saldos_finais):.2f}")
            print(f"Pior saldo final: R$ {min(saldos_finais):.2f}")
            print(f"Média de ajustes: {sum(s['num_ajustes'] for s in simulacoes) / len(simulacoes):.1f}")
            
        # Resumo da aba "Todos"
        saldos_todos = [c['saldo_final'] for c in todos_cenarios if c['saldo_final'] is not None]
        if saldos_todos:
            print(f"\nResumo da aba 'Todos' (todas as combinações):")
            print("=" * 80)
            print(f"Melhor saldo final: R$ {max(saldos_todos):.2f}")
            print(f"Pior saldo final: R$ {min(saldos_todos):.2f}")
            print(f"Média de ajustes: {sum(c['num_ajustes'] for c in todos_cenarios) / len(todos_cenarios):.1f}")
            
            # Análise por limite de lote
            lotes_unicos = sorted(set(c['limite_lote'] for c in todos_cenarios))
            print(f"\nLimites de Lote testados: {lotes_unicos}")
            
            # Análise por volatilidade
            vols_unicos = sorted(set(c['pregoes_volatilidade'] for c in todos_cenarios))
            print(f"Períodos de Volatilidade testados: {vols_unicos}")
        
    except Exception as e:
        print(f"Erro durante a análise: {str(e)}")
    
    finally:
        conn.close()

if __name__ == "__main__":
    main() 